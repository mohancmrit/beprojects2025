import openpyxl

# Load the Excel file with error handling
try:
    file_path = 'C__Users_mohan_OneDrive_Desktop_mega_project_Project_ESA-Backend_updatedExcels_05-02-2025_FN.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    print("Excel file loaded successfully.")
except Exception as e:
    print(f"Error loading the Excel file: {e}")
    exit(1)

# Check if there are any sheets in the workbook
if not workbook.sheetnames:
    print("No sheets found in the Excel file.")
    exit(1)

print(f"Sheets found: {workbook.sheetnames}")

for sheet in workbook.sheetnames:
    print(f'Sheet: {sheet}')
    worksheet = workbook[sheet]
    for row in worksheet.iter_rows(values_only=True):
        print(row)
